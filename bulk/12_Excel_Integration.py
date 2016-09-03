
# coding: utf-8

# <img src="http://hilpisch.com/tpq_logo.png" alt="The Python Quants" width="35%" align="right" border="0"><br>

# # Python for Finance

# **Analyze Big Financial Data**
# 
# O'Reilly (2014)
# 
# Yves Hilpisch

# <img style="border:0px solid grey;" src="http://hilpisch.com/python_for_finance.png" alt="Python for Finance" width="30%" align="left" border="0">

# **Buy the book ** |
# <a href='http://shop.oreilly.com/product/0636920032441.do' target='_blank'>O'Reilly</a> |
# <a href='http://www.amazon.com/Yves-Hilpisch/e/B00JCYHHJM' target='_blank'>Amazon</a>
# 
# **All book codes & IPYNBs** |
# <a href="http://oreilly.quant-platform.com">http://oreilly.quant-platform.com</a>
# 
# **The Python Quants GmbH** | <a href='http://pythonquants.com' target='_blank'>www.pythonquants.com</a>
# 
# **Contact us** | <a href='mailto:analytics@pythonquants.com'>analytics@pythonquants.com</a>

# # Excel Integration

# ## Basic Spreadsheet Interaction

# In[1]:

import numpy as np
import pandas as pd
import xlrd, xlwt
import xlsxwriter
path = 'data/'


# ### Generating Workbooks (xls)

# In[2]:

wb = xlwt.Workbook()


# In[3]:

wb


# In[4]:

wb.add_sheet('first_sheet', cell_overwrite_ok=True)


# In[5]:

wb.get_active_sheet()


# In[6]:

ws_1 = wb.get_sheet(0)
ws_1


# In[7]:

ws_2 = wb.add_sheet('second_sheet')


# In[8]:

data = np.arange(1, 65).reshape((8, 8))


# In[9]:

data


# In[10]:

ws_1.write(0, 0, 100)
  # write 100 in cell "A1"


# In[11]:

for c in range(data.shape[0]):
    for r in range(data.shape[1]):
        ws_1.write(r, c, data[c, r])
        ws_2.write(r, c, data[r, c])


# In[12]:

wb.save(path + 'workbook.xls')


# ### Generating Workbooks (xslx)

# In[13]:

wb = xlsxwriter.Workbook(path + 'workbook.xlsx')


# In[14]:

ws_1 = wb.add_worksheet('first_sheet')
ws_2 = wb.add_worksheet('second_sheet')


# In[15]:

for c in range(data.shape[0]):
    for r in range(data.shape[1]):
        ws_1.write(r, c, data[c, r])
        ws_2.write(r, c, data[r, c])


# In[16]:

wb.close()


# In[17]:

ll $path*


# In[18]:

wb = xlsxwriter.Workbook(path + 'chart.xlsx')
ws = wb.add_worksheet()

# write cumsum of random values in first column
values = np.random.standard_normal(15).cumsum()
ws.write_column('A1', values)

# create a new chart object
chart = wb.add_chart({'type': 'line'})

# add a series to the chart
chart.add_series({'values': '=Sheet1!$A$1:$A$15',
                  'marker': {'type': 'diamond'},})
  # series with markers (here: diamond)

# insert the chart
ws.insert_chart('C1', chart)

wb.close()


# ### Reading from Workbooks

# In[19]:

book = xlrd.open_workbook(path + 'workbook.xlsx')


# In[20]:

book


# In[21]:

book.sheet_names()


# In[22]:

sheet_1 = book.sheet_by_name('first_sheet')
sheet_2 = book.sheet_by_index(1)
sheet_1


# In[23]:

sheet_2.name


# In[24]:

sheet_1.ncols, sheet_1.nrows


# In[25]:

cl = sheet_1.cell(0, 0)
cl.value


# In[26]:

cl.ctype


# In[27]:

sheet_2.row(3)


# In[28]:

sheet_2.col(3)


# In[29]:

sheet_1.col_values(3, start_rowx=3, end_rowx=7)


# In[30]:

sheet_1.row_values(3, start_colx=3, end_colx=7)


# In[31]:

for c in range(sheet_1.ncols):
    for r in range(sheet_1.nrows):
        print '%i' % sheet_1.cell(r, c).value,
    print


# ### Using OpenPyxl

# In[32]:

import openpyxl as oxl


# In[33]:

wb = oxl.Workbook()


# In[34]:

ws = wb.create_sheet(index=0, title='oxl_sheet')


# In[35]:

for c in range(data.shape[0]):
    for r in range(data.shape[1]):
        ws.cell(row=r, column=c).value = data[c, r]
        # creates a Cell object and assigns a value


# In[36]:

wb.save(path + 'oxl_book.xlsx')


# In[37]:

wb = oxl.load_workbook(path + 'oxl_book.xlsx')


# In[38]:

ws = wb.get_active_sheet()


# In[39]:

cell = ws['B4']


# In[40]:

cell.column


# In[41]:

cell.row


# In[42]:

cell.value


# In[43]:

ws['B1':'B4']


# In[44]:

for cell in ws['B1':'B4']:
    print cell[0].value


# In[45]:

ws.range('B1:C4')
  # same as ws['B1':'C4']


# In[46]:

for row in ws.range('B1:C4'):
    for cell in row:
        print cell.value,
    print


# ### Using pandas for Reading and Writing

# In[47]:

df_1 = pd.read_excel(path + 'workbook.xlsx',
                     'first_sheet', header=None)
df_2 = pd.read_excel(path + 'workbook.xlsx',
                     'second_sheet', header=None)


# In[48]:

import string
columns = []
for c in range(data.shape[0]):
    columns.append(string.uppercase[c])
columns


# In[49]:

df_1.columns = columns
df_2.columns = columns


# In[50]:

df_1


# In[51]:

df_2


# In[52]:

df_1.to_excel(path + 'new_book_1.xlsx', 'my_sheet')


# In[53]:

wbn = xlrd.open_workbook(path + 'new_book_1.xlsx')


# In[54]:

wbn.sheet_names()


# In[55]:

wbw = pd.ExcelWriter(path + 'new_book_2.xlsx')
df_1.to_excel(wbw, 'first_sheet')
df_2.to_excel(wbw, 'second_sheet')
wbw.save()


# In[56]:

wbn = xlrd.open_workbook(path + 'new_book_2.xlsx')


# In[57]:

wbn.sheet_names()


# In[58]:

data = np.random.rand(20, 100000)


# In[59]:

data.nbytes


# In[60]:

df = pd.DataFrame(data)


# In[61]:

get_ipython().magic("time df.to_excel(path + 'data.xlsx', 'data_sheet')")


# In[62]:

get_ipython().magic("time np.save(path + 'data', data)")


# In[63]:

ll $path*


# In[64]:

get_ipython().magic("time df = pd.read_excel(path + 'data.xlsx', 'data_sheet')")


# In[65]:

get_ipython().magic("time data = np.load(path + 'data.npy')")


# In[66]:

data, df = 0.0, 0.0
get_ipython().system('rm $path*')


# ## Scripting Excel with Python

# ### Installing DataNitro

# ### Working with DataNitro

# #### Scripting with DataNitro

# #### Plotting with DataNitro

# #### User Defined Functions

# ## xlwings

# ## Conclusions

# ## Further Reading

# <img src="http://hilpisch.com/tpq_logo.png" alt="The Python Quants" width="35%" align="right" border="0"><br>
# 
# <a href="http://www.pythonquants.com" target="_blank">www.pythonquants.com</a> | <a href="http://twitter.com/dyjh" target="_blank">@dyjh</a>
# 
# <a href="mailto:analytics@pythonquants.com">analytics@pythonquants.com</a>
# 
# **Python Quant Platform** |
# <a href="http://oreilly.quant-platform.com">http://oreilly.quant-platform.com</a>
# 
# **Derivatives Analytics with Python** |
# <a href="http://www.derivatives-analytics-with-python.com" target="_blank">Derivatives Analytics @ Wiley Finance</a>
# 
# **Python for Finance** |
# <a href="http://shop.oreilly.com/product/0636920032441.do" target="_blank">Python for Finance @ O'Reilly</a>
